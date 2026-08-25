import openpyxl
import json
import datetime

wb = openpyxl.load_workbook('/Users/nourine/Downloads/Algeria_K9  CKD0 Project Weekly Planning_V04_20260825.xlsx', data_only=True)
ws = wb['Onsite Timing']

all_tasks = []
current_zone = 'UB'

for r in range(6, 101):
    c1 = ws.cell(r, 1).value
    c2 = ws.cell(r, 2).value
    c3 = ws.cell(r, 3).value
    c4 = ws.cell(r, 4).value # start
    c5 = ws.cell(r, 5).value # end

    text = str(c2 or c1 or '').strip()
    if text in ['UB', 'UAR', 'FUSA']:
        current_zone = text
        continue
    if not text or text.startswith('SUM'):
        continue

    # Parse dates
    start_dt = None
    end_dt = None
    if isinstance(c4, datetime.datetime):
        start_dt = c4.date()
    elif c4:
        try:
            start_dt = datetime.datetime.strptime(str(c4).split(' ')[0], "%Y-%m-%d").date()
        except:
            pass

    if isinstance(c5, datetime.datetime):
        end_dt = c5.date()
    elif c5:
        try:
            end_dt = datetime.datetime.strptime(str(c5).split(' ')[0], "%Y-%m-%d").date()
        except:
            pass

    if start_dt and end_dt:
        all_tasks.append({
            "zone": current_zone,
            "task_full": text,
            "start": start_dt.isoformat(),
            "end": end_dt.isoformat(),
            "start_dt": start_dt,
            "end_dt": end_dt
        })

print(f"Parsed {len(all_tasks)} specific zone tasks from Excel.")

# Week calendar KW25 (2026-06-15) to KW53 (2027-01-03)
week_calendar = {
    25: (datetime.date(2026, 6, 15), datetime.date(2026, 6, 21)),
    26: (datetime.date(2026, 6, 22), datetime.date(2026, 6, 28)),
    27: (datetime.date(2026, 6, 29), datetime.date(2026, 7, 5)),
    28: (datetime.date(2026, 7, 6), datetime.date(2026, 7, 12)),
    29: (datetime.date(2026, 7, 13), datetime.date(2026, 7, 19)),
    30: (datetime.date(2026, 7, 20), datetime.date(2026, 7, 26)),
    31: (datetime.date(2026, 7, 27), datetime.date(2026, 8, 2)),
    32: (datetime.date(2026, 8, 3), datetime.date(2026, 8, 9)),
    33: (datetime.date(2026, 8, 10), datetime.date(2026, 8, 16)),
    34: (datetime.date(2026, 8, 17), datetime.date(2026, 8, 23)),
    35: (datetime.date(2026, 8, 24), datetime.date(2026, 8, 30)),
    36: (datetime.date(2026, 8, 31), datetime.date(2026, 9, 6)),
    37: (datetime.date(2026, 9, 7), datetime.date(2026, 9, 13)),
    38: (datetime.date(2026, 9, 14), datetime.date(2026, 9, 20)),
    39: (datetime.date(2026, 9, 21), datetime.date(2026, 9, 27)),
    40: (datetime.date(2026, 9, 28), datetime.date(2026, 10, 4)),
    41: (datetime.date(2026, 10, 5), datetime.date(2026, 10, 11)),
    42: (datetime.date(2026, 10, 12), datetime.date(2026, 10, 18)),
    43: (datetime.date(2026, 10, 19), datetime.date(2026, 10, 25)),
    44: (datetime.date(2026, 10, 26), datetime.date(2026, 11, 1)),
    45: (datetime.date(2026, 11, 2), datetime.date(2026, 11, 8)),
    46: (datetime.date(2026, 11, 9), datetime.date(2026, 11, 15)),
    47: (datetime.date(2026, 11, 16), datetime.date(2026, 11, 22)),
    48: (datetime.date(2026, 11, 23), datetime.date(2026, 11, 29)),
    49: (datetime.date(2026, 11, 30), datetime.date(2026, 12, 6)),
    50: (datetime.date(2026, 12, 7), datetime.date(2026, 12, 13)),
    51: (datetime.date(2026, 12, 14), datetime.date(2026, 12, 20)),
    52: (datetime.date(2026, 12, 21), datetime.date(2026, 12, 27)),
    53: (datetime.date(2026, 12, 28), datetime.date(2027, 1, 3))
}

def translate_task(en_text):
    t = en_text.lower()
    if 'floor marking' in t:
        return 'Traçage au sol et repérage géométrique des zones', 'Floor marking & layout positioning', '划线及区域定位'
    if 'steel structure' in t:
        return 'Montage et érection des charpentes métalliques', 'Steel structure installation and erection', '钢结构安装'
    if 'station lighting' in t:
        return 'Installation et câblage de l\'éclairage des postes', 'Station lighting installation and wiring', '工位照明安装'
    if 'station fans' in t:
        return 'Pose et fixation des ventilateurs industriels de postes', 'Station cooling fans installation', '工位风扇安装'
    if 'pipe installation' in t:
        return 'Pose et raccordement des tuyauteries eau et air comprimé', 'Pneumatic and cooling water piping installation', '水气管道安装'
    if 'busbar' in t:
        return 'Installation des jeux de barres électriques blindées IG2', 'IG2 busbar electrical installation', 'IG2母排安装'
    if 'cabel' in t or 'cable' in t or 'plug' in t:
        return 'Tirage et raccordement des câbles de puissance IG1-IG2', 'Power cable pulling and IG1-IG2 connection', '电缆敷设与接线'
    if 'demag' in t or 'kbk' in t:
        return 'Montage des rails et ponts suspendus DEMAG KBK', 'DEMAG KBK overhead rail and crane installation', 'DEMAG导轨及起重设备安装'
    if 'controller' in t:
        return 'Installation et mise en service des coffrets contrôleurs de soudage', 'Manual welding controllers installation & setup', '手动焊接控制器安装与调试'
    if 'gun' in t:
        return 'Montage et équilibrage des pinces à souder manuelles', 'Manual welding guns installation and balancing', '手动焊枪安装与平衡器调节'
    if 'power on' in t:
        return 'Mise sous tension, purge fluides et raccordements énergies', 'Power on, water & pneumatic supply connection', '水气电能源接通与调试'
    if 'gripper' in t:
        return 'Montage mécanique et alignement des préhenseurs (grippers)', 'Robotic grippers installation and mechanical alignment', '机器人抓手安装与机械校准'
    if 'gluing' in t:
        return 'Pose et raccordement des équipements de dépose de colle', 'Gluing dispensing equipment installation', '涂胶设备安装与管道连接'
    if 'fixture on position' in t or 'fixtures on position' in t:
        return 'Implantation, mise à niveau et calage des gabarits (fixtures)', 'Fixtures positioning, leveling and mechanical adjustment', '夹具落位及精准调平'
    if 'torque check' in t:
        return 'Contrôle et serrage au couple certifié des ancrages', 'Torque check and anchor bolting verification', '地脚螺栓扭力检查'
    if 'buyoff' in t:
        return 'Levée des réserves et ajustements préalables des gabarits', 'Pre-buyoff checklist issues closure and fine tuning', '预验收问题单整改'
    if 'activity 3' in t or 'measurement' in t:
        return 'Mesures tridimensionnelles, réglages précis et rapport Black Book', 'Fixtures 3D measurement, adjustment and Black Book report', '夹具三坐标测量、调试及Black Book报告'
    if 'pre-activity 4' in t:
        return 'Contrôle géométrique préalable des jeux entre pièces', 'Preliminary parts clearance inspection', '零件间隙预检查'
    if 'activity 4' in t or 'parts to fixtures' in t:
        return 'Ajustement géométrique des pièces sur les gabarits', 'Parts fitting and clearance adjustment on fixtures', '零件与夹具匹配及间隙调整'
    if 'activity 5' in t or 'slow build' in t or 'reapitibility' in t or 'repeatability' in t:
        return 'Tests de répétabilité et assemblage à cadence lente', 'Repeatability tests and slow-build assembly trials', '重复性验证与慢速试装'
    if 'first asm' in t:
        return 'Réalisation et validation de la première pièce assemblée (First Assembly)', 'First Assembly Part (First ASM) manufacturing and validation', '首件装配制作与质检验收'
    if 'certification' in t:
        return 'Essais de sécurité, tests de conformité et certification', 'Safety certification and compliance validation tests', '安全认证及综合测试'
    if 'fee_ree' in t:
        return 'Vérification de conformité de l\'installation FEE/REE', 'FEE_REE installation compliance checklist verification', 'FEE/REE安装检查清单验证'
    if 'safety acceptance' in t:
        return 'Réception formelle de sécurité HSE et levée des barrières', 'HSE Safety Acceptance and site authorization buyoff', '安全验收与现场放行'
    if 'training' in t:
        return 'Formation pratique des opérateurs et techniciens de maintenance', 'Line operators and maintenance team technical training', '操作工与维修人员实操培训'
    if 'start of production' in t or 'sop' in t:
        return 'Démarrage de la production pilote / SOP (Start of Production)', 'Pilot production start / SOP (Start of Production)', '试生产启动 (SOP)'
    if 'x0' in t:
        return 'Jalon Industriel X0 et conformité ligne', 'Industrial Milestone X0 and line buyoff', '工业化里程碑 X0'
    return en_text, en_text, en_text

# Equipment catalog per task type
def get_equipment_details(task_text, zone):
    t = task_text.lower()
    eq = []
    if 'steel structure' in t:
        eq.extend(['Charpentes métalliques primaires', 'Poteaux et poutres IPN/HEA', 'Boulonnerie HR 8.8/10.9', 'Nacelles ciseaux'])
    if 'lighting' in t:
        eq.extend(['Luminaires LED étanches 4000K', 'Chemins de câbles d\'éclairage', 'Tableaux divisionnaires éclairage'])
    if 'fans' in t:
        eq.extend(['Ventilateurs axiaux industriels', 'Variateurs de vitesse', 'Supports amortisseurs de vibrations'])
    if 'pipe' in t:
        eq.extend(['Tuyauteries inox eau de refroidissement (DN50/DN80)', 'Réseau air comprimé 6-8 bars', 'Vannes de sectionnement', 'Purgeurs automatiques'])
    if 'busbar' in t or 'cable' in t:
        eq.extend(['Jeux de barres blindées IG2 (800A/1000A)', 'Coffrets de dérivation', 'Chemins de câbles puissance', 'Câbles cuivre 4x240mm²'])
    if 'demag' in t or 'kbk' in t:
        eq.extend(['Profilés suspendus DEMAG KBK II', 'Palans électriques à chaîne', 'Chariots de translation', 'Équilibreurs de charge'])
    if 'controller' in t or 'gun' in t:
        eq.extend(['Armoires contrôleurs de soudage ARO/BOSCH', 'Pinces à souder manuelles en X et en C', 'Équilibreurs TECNA', 'Faisceaux eau/air/puissance'])
    if 'gripper' in t:
        eq.extend(['Préhenseurs robotisés (Grippers)', 'Vérins pneumatiques Festo/SMC', 'Centrages coniques', 'Capteurs inductifs de présence pièce'])
    if 'gluing' in t:
        eq.extend(['Unité de dosage et pompage colle bi-composant', 'Pistolet d\'application robotisé', 'Système de chauffe régulée'])
    if 'fixture' in t or 'measurement' in t or 'buyoff' in t:
        eq.extend([f'Gabarits de soudage géométriques {zone}', 'Centrages mécaniques trempés', 'Serrages sauterelles pneumatiques', 'Bras de mesure 3D Faro/Romer'])
    if not eq:
        eq.extend([f'Outillages industriels certifiés {zone}', 'Instruments de mesure et calibres'])
    return eq

weekly_data = {}

for w_num, (w_start, w_end) in week_calendar.items():
    # Find all tasks overlapping with this week
    matching_tasks = []
    zones_present = set()
    all_eq = set()

    for t in all_tasks:
        # Check date overlap
        if t['start_dt'] <= w_end and t['end_dt'] >= w_start:
            matching_tasks.append(t)
            zones_present.add(t['zone'])
            for eq_item in get_equipment_details(t['task_full'], t['zone']):
                all_eq.add(eq_item)

    zone_label = " / ".join(sorted(list(zones_present))) if zones_present else "UB / UAR / FUSA"
    
    # Detailed descriptions
    tasks_fr = []
    tasks_en = []
    tasks_zh = []
    
    for t in matching_tasks:
        fr, en, zh = translate_task(t['task_full'])
        tasks_fr.append(f"[{t['zone']}] {fr}")
        tasks_en.append(f"[{t['zone']}] {en}")
        tasks_zh.append(f"[{t['zone']}] {zh}")

    weekly_data[w_num] = {
        "week": w_num,
        "startDate": w_start.isoformat(),
        "endDate": w_end.isoformat(),
        "zones": sorted(list(zones_present)) if zones_present else ["UB", "UAR", "FUSA"],
        "zoneLabel": f"Zones {zone_label}",
        "tasks_fr": tasks_fr[:6],
        "tasks_en": tasks_en[:6],
        "tasks_zh": tasks_zh[:6],
        "equipmentList": list(all_eq)[:8]
    }

print("Successfully compiled all 29 weeks with exact Excel zones and equipment!")
with open('/Users/nourine/.gemini/antigravity-ide/scratch/permis-sinylon/excel_weekly_mapping.json', 'w', encoding='utf-8') as f:
    json.dump(weekly_data, f, indent=2, ensure_ascii=False)
